import os
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import NiveisDeRuido
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import NiveisDeRuido
from .serializers import NiveisDeRuidoSerializer
from rest_framework import viewsets
from datetime import datetime, timedelta
from django.http import HttpResponse
import logging
from django.db.models import Q
from django.http import JsonResponse
from .models import NiveisDeRuido
from django.db.models import F
import matplotlib.dates as mdates
from io import BytesIO
import matplotlib.pyplot as plt
import plotly.express as px
import pandas as pd
import plotly.graph_objs as go
from plotly.offline import plot
import matplotlib.pyplot as plt
import base64
from io import BytesIO
import numpy as np
from django.shortcuts import redirect
from django.contrib.auth import logout
from io import StringIO
import csv
from fpdf import FPDF
import tempfile 
from PIL import Image
import matplotlib
matplotlib.use('Agg')  
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
from openpyxl.styles import PatternFill
import base64
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from collections import defaultdict


#FUNCIONANDO!
@login_required
def dashboard(request):
    total_medicoes = NiveisDeRuido.objects.count() 

    niveis_atividade = NiveisDeRuido.objects.filter(status=True)
    
    tempo_atividade_total = timedelta()
    for nivel in niveis_atividade:
        data_hora = datetime.combine(nivel.data, nivel.hora)
        tempo_atividade_total += timedelta(hours=1)  
    horas_atividade = tempo_atividade_total.total_seconds() / 3600  

    try:
        ultima_medicao_registro = NiveisDeRuido.objects.latest('data', 'hora') 
        ultima_medicao_decibel = ultima_medicao_registro.decibeis
        ultima_medicao_status = ultima_medicao_registro.status
    except NiveisDeRuido.DoesNotExist:
        ultima_medicao_decibel = None
        ultima_medicao_status = False 

    if ultima_medicao_status:
        status_sensor = "ON"
    else:
        status_sensor = "OFF"
    
    nome_sensor = "Sensor 1"
    
    return render(request, 'frontend/dashboard.html', {
        'total_medicoes': total_medicoes,
        'horas_atividade': round(horas_atividade, 2),  
        'ultima_medicao': ultima_medicao_decibel,  
        'status_sensor': status_sensor,
        'nome_sensor': nome_sensor,
        
    })









#FUNCIONANDO!
@login_required
def relatorio(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return render(request, 'frontend/relatorio.html', {'usuario': request.user})








#Rapaz hora funciona e caso não receber o Post aper o boot do esp
class ReceberRuido(APIView):
    permission_classes = [AllowAny]  

    def post(self, request):
        try:
            dados = request.data
            
            data = datetime.strptime(dados['data'], '%Y-%m-%d').date()
            hora = datetime.strptime(dados['hora'], '%H:%M:%S').time()
            decibeis = dados['decibeis']
            limite = dados['limite']
            status = dados['status']

            novo_nivel = NiveisDeRuido(
                data=data,
                hora=hora,
                decibeis=decibeis,
                limite=limite,
                status=status
            )
            
            novo_nivel.save()

            return Response({
                "status": "ok",
                "message": "Nível de ruído registrado com sucesso",
                "data": dados
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                "status": "error",
                "message": f"Erro ao processar os dados: {str(e)}"
            }, status=status.HTTP_400_BAD_REQUEST)


class NiveisDeRuidoViewSet(viewsets.ModelViewSet):
    queryset = NiveisDeRuido.objects.all()
    serializer_class = NiveisDeRuidoSerializer



#FUNCIONANDO!
#A desgraça mais difícil de todo esse trabalho
#Um dia não funcionava, agora ela funciona
#usei o dicionário para exibir no HTML mas  dava erro e não exibia nada e continuava assim
#até que teve a ideia de usar a messma .data para o HTML e essa coisa funcionou.
@login_required
def historico_medicoes(request):
    periodo_dia = request.GET.get('periodo-dia')
    perturbacao = request.GET.get('perturbacao')
    mes_registro = request.GET.get('mes-registro')

    niveis = NiveisDeRuido.objects.all()

    if periodo_dia:
        if periodo_dia == 'manha':
            niveis = niveis.filter(hora__lt='12:00:00')
        elif periodo_dia == 'tarde':
            niveis = niveis.filter(hora__gte='12:00:00', hora__lt='18:00:00')
        elif periodo_dia == 'noite':
            niveis = niveis.filter(hora__gte='18:00:00')

    if perturbacao:
        if perturbacao == 'sim':
            niveis = niveis.filter(status=True)
        elif perturbacao == 'nao':
            niveis = niveis.filter(status=False)

    if mes_registro:
        niveis = niveis.filter(data__month=mes_registro)

    context = {
        'niveis': niveis,
        'periodo_dia': periodo_dia,
        'perturbacao': perturbacao,
    }
    return render(request, 'relatorio.html', context)




#ESTE JA ESTÁ FUNCIONANDO


def grafico(request):
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    try:
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Datas inválidas fornecidas.", status=400)

    
    niveis = NiveisDeRuido.objects.all()


    if data_inicio and data_fim:
        niveis = niveis.filter(data__range=(data_inicio, data_fim))

   
    print(f"Buscando dados entre {data_inicio} e {data_fim}")
    

    date_dict = defaultdict(list)
    
    for nivel in niveis:
        if nivel.data and nivel.hora:
            combined_date = datetime.combine(nivel.data, nivel.hora)
            date_dict[combined_date].append(nivel.decibeis)

    dates = list(date_dict.keys())
    decibeis = [sum(decibels)/len(decibels) for decibels in date_dict.values()]

    if not decibeis:
        return HttpResponse("Nenhum dado disponível para o período selecionado.", status=404)

    print(f"Datas a serem exibidas no gráfico: {dates}")
    print(f"Decibéis correspondentes: {decibeis}")

    sorted_data = sorted(zip(dates, decibeis), key=lambda x: x[0])
    dates, decibeis = zip(*sorted_data)

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(dates, decibeis, color='skyblue', width=0.12)

    ax.set_xlabel('Data e Hora')
    ax.set_ylabel('Nível de Ruído (dB)')
    ax.set_title('Gráfico de Níveis de Ruído ao Longo do Tempo')

    plt.xticks(rotation=45, ha='right')

    ax.xaxis.set_major_locator(mdates.AutoDateLocator()) 
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M:%S'))

  
    ax.set_ylim(0, max(decibeis) * 1.1)
    plt.tight_layout()

 
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    img_data = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close(fig)


    return render(request, 'frontend/grafico.html', {'imagem_grafico': img_data})




def logout_view(request):
    logout(request) 
    return redirect('/')




#FUNCIONANDO!

def nome_mes_sem_acentuacao(mes_num):
    meses = [
        "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    return meses[mes_num - 1] if 1 <= mes_num <= 12 else 'Mes invalido'
from django.core.exceptions import ValidationError

def gerar_relatorio(request):
    periodo_dia = request.GET.get('periodo-dia')
    perturbacao = request.GET.get('perturbacao')
    mes_registro = request.GET.get('mes-registro')

    niveis = NiveisDeRuido.objects.all()

    if periodo_dia:
        if periodo_dia == 'manha':
            niveis = niveis.filter(hora__lt='12:00:00')
        elif periodo_dia == 'tarde':
            niveis = niveis.filter(hora__gte='12:00:00', hora__lt='18:00:00')
        elif periodo_dia == 'noite':
            niveis = niveis.filter(hora__gte='18:00:00')
        else:
            raise ValidationError("Período do dia inválido")

    if perturbacao:
        if perturbacao == 'sim':
            niveis = niveis.filter(status=True)
        elif perturbacao == 'nao':
            niveis = niveis.filter(status=False)
        else:
            raise ValidationError("Valor de perturbação inválido")

    if mes_registro:
        try:
            mes_registro = int(mes_registro)
            niveis = niveis.filter(data__month=mes_registro)
        except ValueError:
            raise ValidationError("Mês de registro inválido")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Níveis de Ruído"

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    ws.append(['Relatório Técnico de Níveis de Ruído'])
    ws.append(['Este relatório contém dados sobre os níveis de ruído coletados pela plataforma EchoWarden.'])
    ws.append(['A plataforma EchoWarden é um sistema de monitoramento ambiental projetado para registrar e analisar os níveis de ruído em áreas urbanas e industriais.'])
    ws.append(['Ele utiliza sensores acústicos distribuídos para coletar dados em tempo real, permitindo uma análise detalhada das condições de ruído e seu impacto na saúde pública e qualidade de vida.'])
    ws.append([f'Relatório gerado em: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    ws.append(['Fonte de dados: Banco de Dados - Modelo NiveisDeRuido'])
    ws.append([''])
    
    ws.append(['Campos do Relatório:'])
    ws.append(['- "Data": Data e hora em que o nível de ruído foi registrado. Essa informação é crucial para determinar a tendência temporal dos níveis de ruído e identificar padrões.'])
    ws.append(['- "Nível de Ruído (dB)": Representa o nível de pressão sonora medido em decibéis (dB), uma unidade logarítmica que expressa a intensidade do som. Níveis mais altos podem indicar poluição sonora e impacto ambiental.'])
    ws.append(['- "Status de Perturbação": Indica se o nível de ruído registrado ultrapassou o limite tolerável para o ambiente em questão. A presença de um "status de perturbação" sugere que os níveis de ruído podem ser prejudiciais.'])
    ws.append(['- "Mês": Mês em que o dado foi registrado. Isso ajuda a observar variações sazonais ou tendências anuais no comportamento do ruído.'])
    ws.append(['- "Período do dia": Define a faixa horária em que o dado foi coletado, ajudando a analisar as variações nos níveis de ruído ao longo do dia.'])
    ws.append([''])
    
    ws.append([f'Filtro - Período do dia: {periodo_dia if periodo_dia else "Todos os períodos"}'])
    ws.append([f'Filtro - Status de perturbação: {perturbacao if perturbacao else "Todos os status"}'])
    ws.append([f'Filtro - Mês de registro: {mes_registro if mes_registro else "Todos os meses"}'])
    ws.append([''])

    ws.append(['Descrição do Status de Perturbação:'])
    ws.append(['- "Perturbado" (status=1): Indica que o nível de ruído registrado ultrapassou o limite estabelecido, podendo causar impactos na saúde humana e no bem-estar social.'])
    ws.append(['- "Normal" (status=0): Indica que o nível de ruído está dentro dos parâmetros aceitáveis para o ambiente e não é considerado perturbador.'])
    ws.append([''])
    
    ws.append(['Limites de Nível de Ruído:'])
    ws.append(['- Os limites de ruído considerados seguros podem variar de acordo com o local (urbano, industrial, residencial) e a legislação local. Em áreas residenciais, por exemplo, níveis superiores a 55 dB durante a noite são geralmente considerados prejudiciais.'])
    ws.append(['- O monitoramento contínuo de ruído é fundamental para identificar fontes de poluição sonora e adotar medidas corretivas, como o controle de tráfego ou regulamentação de horários de funcionamento de indústrias.'])
    ws.append([''])
    
    ws.append(['Data', 'Nível de Ruído (dB)', 'Status de Perturbação', 'Mês', 'Período do dia'])
    
    for cell in ws[1]:
        cell.fill = header_fill  

    for nivel in niveis:
        if nivel.data:
            data_formatada = nivel.data.strftime("%Y-%m-%d %H:%M:%S")
            mes = nivel.data.month
        else:
            data_formatada = 'Data não disponível'
            mes = 'Desconhecido'  
        #Essa porcaria so entende assim porque lá no banco está como true ou false
        #então ela não pega os dada e fica aparecendo "none"
        
        #lembrar que if data > month_excedente não aparee
        status_texto = 'Perturbado' if nivel.status else 'Normal'

        row = [data_formatada, nivel.decibeis, status_texto, mes, nivel.hora]
        ws.append(row)

        last_row = ws.max_row

        if nivel.decibeis > nivel.limite:
            for col in range(1, len(row) + 1): 
                ws.cell(row=last_row, column=col).fill = red_fill  

    ws.append([''])
    ws.append(['Conclusão:'])
    ws.append(['Este relatório fornece uma visão detalhada sobre os níveis de ruído em áreas monitoradas pela plataforma EchoWarden, permitindo a identificação de períodos críticos de poluição sonora e o acompanhamento da eficácia de medidas corretivas, caso necessário.'])
    ws.append(['Com base nas medições coletadas e nos filtros aplicados, é possível observar padrões de variação nos níveis de ruído ao longo do tempo, bem como a relação entre o status de perturbação e os horários de maior incidência de níveis elevados de ruído.'])
    ws.append(['A coleta contínua e a análise desses dados são fundamentais para a implementação de políticas públicas eficazes de controle da poluição sonora e para a melhoria da qualidade de vida nas comunidades.'])
    ws.append(['Medidas corretivas podem incluir ajustes no planejamento urbano, como a regulamentação de horários de funcionamento de atividades industriais e comerciais, a implantação de zonas silenciosas em áreas residenciais, e o incentivo ao uso de tecnologias que reduzam a emissão de ruído.'])
    ws.append(['Além disso, a conscientização pública sobre os impactos da poluição sonora na saúde humana deve ser uma prioridade, com campanhas educativas sobre limites de exposição ao ruído e práticas adequadas para minimizar os efeitos negativos da exposição contínua ao som excessivo.'])
    ws.append(['A plataforma EchoWarden, ao fornecer dados em tempo real sobre os níveis de ruído, pode ser um instrumento valioso para órgãos de regulamentação, comunidades locais e pesquisadores que buscam entender e combater a poluição sonora.'])
    ws.append(['A continuidade do monitoramento será essencial para garantir que as intervenções adotadas sejam eficazes na melhoria das condições acústicas das áreas monitoradas.'])
    ws.append([''])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(file_stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_niveis_ruido.xlsx"'

    return response





#FUNCIONANDO!

def relatoriofinal(request):
    periodo_dia = request.GET.get('periodo-dia')
    perturbacao = request.GET.get('perturbacao')
    mes_registro = request.GET.get('mes-registro')

    niveis = NiveisDeRuido.objects.all()

    if periodo_dia:
        if periodo_dia == 'manha':
            niveis = niveis.filter(hora__lt='12:00:00')
        elif periodo_dia == 'tarde':
            niveis = niveis.filter(hora__gte='12:00:00', hora__lt='18:00:00')
        elif periodo_dia == 'noite':
            niveis = niveis.filter(hora__gte='18:00:00')
        else:
            raise ValidationError("Período do dia inválido")

    if perturbacao:
        if perturbacao == 'sim':
            niveis = niveis.filter(status=True)
        elif perturbacao == 'nao':
            niveis = niveis.filter(status=False)
        else:
            raise ValidationError("Valor de perturbacao inválido")

    if mes_registro:
        try:
            mes_registro = int(mes_registro)  #TEM QUE CONVERTER PARA INTERIROR PARA FUNCIONAR
            if not (1 <= mes_registro <= 12):
                raise ValidationError("Mês de registro inválido")
            niveis = niveis.filter(data__month=mes_registro)
        except ValueError:
            raise ValidationError("Mês de registro inválido")

    context = {
        'niveis': niveis,
        'periodo_dia': periodo_dia,
        'perturbacao': perturbacao,
        'mes_registro': mes_registro
    }

    return render(request, 'frontend/relatorio.html', context)
